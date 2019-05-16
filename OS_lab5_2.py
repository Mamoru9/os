import xlrd, openpyxl
from openpyxl.styles import PatternFill
from array import *

#Структура процесса
class Flow(object):
    def __init__(self, name=None, time=2000, priority=1000, time_start=3000, time_w8=0):
        self.name = name
        self.time = time
        self.priority = priority
        self.time_start = time_start
        self.time_w8 = time_w8

    def null_time_w8(self):
        while self.time_w8 !=0:
            self.time_w8 -= 1

    def inc_time_w8(self):
        self.time_w8 += 1

#Функция считающее время ожидание
def expecttion (data, obj):
    obj.null_time_w8()
    for i in range(30):
        if i == obj.time_start:
            while data[i] != 0:
                obj.inc_time_w8()
                i += 1
            break

def fill_cvant_2 (flag, obj1, obj2, new_flag):
    if new_flag == 1:
        if flag == 0:
            temp = obj1.time_w8
            for j in range(int(obj1.time + obj1.time_start + obj1.time_w8)):
                if j< obj1.time_start:
                    continue
                while temp:
                    sheet.cell(row =5, column=j+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    temp -=1
                    j += 1
                if data[j] != 0:
                    continue
                sheet.cell(row =5, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
                data[j] += 1
            expecttion (data, obj2)
            temp = obj2.time_w8
            for i in range(int(obj2.time + obj2.time_start + obj2.time_w8)):
                if i< obj2.time_start:
                    continue
                while temp:
                    sheet.cell(row =6, column=i+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    temp -=1
                    i += 1
                if data[j] != 0:
                    continue
                sheet.cell(row =6, column=i+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
                data[j] += 1

        elif flag == 1:
            temp = obj2.time_w8            
            for j in range(int(obj2.time + obj2.time_start + obj2.time_w8)):
                if j< obj2.time_start:
                    continue
                while temp:
                    sheet.cell(row =6, column=j+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    temp -=1
                    j += 1
                if data[j] != 0:
                    continue
                sheet.cell(row =6, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
                data[j] += 1
            expecttion (data, obj1)
            temp = obj1.time_w8
            for i in range(int(obj1.time + obj1.time_start + obj1.time_w8)):
                if i< obj1.time_start:
                    continue
                while temp:
                    sheet.cell(row =5, column=i+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    temp -=1
                    i += 1
                if data[j] != 0:
                    continue
                sheet.cell(row =5, column=i+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
                data[i] += 1
    elif new_flag == 2:
        if flag == 0:
            temp = obj1.time_w8
            for j in range(int(obj1.time_start + obj1.time_w8 + obj1.time)):
                if j< obj1.time_start:
                    continue
                while temp:
                    sheet.cell(row =4, column=j+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    temp -=1
                    j += 1
                if data[j] != 0:
                    continue
                sheet.cell(row =4, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
                data[j] += 1
            expecttion (data, obj2)
            temp = obj2.time_w8
            for i in range(int(obj2.time + obj2.time_start + obj2.time_w8)):
                if i< obj2.time_start:
                    continue
                while temp:
                    sheet.cell(row =6, column=i+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    temp -=1
                    i += 1
                if data[i] != 0:
                    continue
                sheet.cell(row =6, column=i+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
                data[i] += 1

        elif flag == 1:
            temp = obj2.time_w8            
            for j in range(int(obj2.time + obj2.time_start + obj2.time_w8)):
                if j< obj2.time_start:
                    continue
                while temp:
                    sheet.cell(row =6, column=j+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    temp -=1
                    j += 1
                if data[j] != 0:
                    continue
                sheet.cell(row =6, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
                data[j] += 1
            expecttion (data, obj1)
            temp = obj1.time_w8
            for j in range(int(obj1.time + obj1.time_start + obj1.time_w8)):
                if j< obj1.time_start:
                    continue
                while temp:
                    sheet.cell(row =4, column=j+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    temp -=1
                    j += 1
                if data[j] != 0:
                    continue
                sheet.cell(row =4, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
                data[j] += 1
    elif new_flag == 3:
        if flag == 0:
            temp = obj1.time_w8
            for j in range(int(obj1.time + obj1.time_start + obj1.time_w8)):
                if j< obj1.time_start:
                    continue
                while temp:
                    sheet.cell(row =4, column=j+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    temp -=1
                    j += 1
                if data[j] != 0:
                    continue
                sheet.cell(row =4, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
                data[j] += 1
            expecttion (data, obj2)
            temp = obj2.time_w8  
            for j in range(int(obj2.time + obj2.time_start + obj2.time_w8)):
                if j< obj2.time_start:
                    continue
                while temp:
                    sheet.cell(row =5, column=j+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    temp -=1
                    j += 1
                if data[j] != 0:
                    continue
                sheet.cell(row =5, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
                data[j] += 1
        elif flag == 1:
            temp = obj2.time_w8            
            for j in range(int(obj2.time + obj2.time_start + obj2.time_w8)):
                if j< obj2.time_start:
                    continue
                while temp:
                    sheet.cell(row =5, column=j+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    temp -=1
                    j += 1
                if data[j] != 0:
                    continue
                sheet.cell(row =5, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
                data[j] += 1
            expecttion (data, obj1)
            temp = obj1.time_w8  
            for j in range(int(obj1.time + obj1.time_start + obj1.time_w8)):
                if j< obj1.time_start:
                    continue
                while temp:
                    sheet.cell(row =4, column=j+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    temp -=1
                    j += 1
                if data[j] != 0:
                    continue
                sheet.cell(row =4, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
                data[j] += 1

def fill_cvant_1 (flag, obj1, obj2, obj3):
    if flag == 0:
        temp = obj1.time_w8
        for j in range(int(obj1.time + obj1.time_start + obj1.time_w8)):
            if j< obj1.time_start:
                continue
            while temp:
                sheet.cell(row =4, column=j+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                temp -=1
                j += 1
            if data[j] != 0:
                continue
            sheet.cell(row =4, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
            data[j] += 1
        expecttion (data, obj2)
        expecttion (data, obj3)
        min_start = obj2.time_start
        if min_start > obj3.time_start:
            flag = 1
        fill_cvant_2(flag, obj2, obj3, 1)
    elif flag == 1:
        temp = obj2.time_w8
        
        for j in range(int(obj2.time + obj2.time_start + obj2.time_w8)):
            if j< obj2.time_start:
                continue
            while temp:
                sheet.cell(row =5, column=j+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                temp -=1
                j += 1
            if data[j] != 0:
                continue
            sheet.cell(row =5, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
            data[j] += 1
        expecttion (data, obj1)
        expecttion (data, obj3)
        flag = 0
        min_start = obj1.time_start
        if min_start > obj3.time_start:
            flag = 1
        fill_cvant_2(flag, obj1, obj3, 2)
    elif flag == 2:
        temp = obj3.time_w8
        for j in range(int(obj3.time + obj3.time_start + obj3.time_w8)):
            if j< obj3.time_start:
                continue
            while temp:
                sheet.cell(row =6, column=j+4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                temp -=1
                j += 1
            if data[j] != 0:
                continue
            sheet.cell(row =6, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
            data[j] += 1
        expecttion (data, obj1)
        expecttion (data, obj2)
        flag = 0
        min_start = obj1.time_start
        if min_start > obj2.time_start:
            flag = 1
        fill_cvant_2(flag, obj1, obj2, 3)

    


#Заполнение структур процессов
sum_struct = 0
rb = xlrd.open_workbook('text.xlsx') #Открытие файла на чтение
sheet = rb.sheet_by_index(0)
for rownum in range(sheet.nrows):
    row = sheet.row_values(rownum)
    i = 0
    for c_el in row:
        if i == 0:
            tmp_name = c_el
        elif i == 1:
            tmp_time = c_el
        elif i == 2:
            tmp_priority = c_el
        elif i == 3:
            tmp_time_start = c_el
        i += 1
    if sum_struct == 0:
        first = Flow(tmp_name, tmp_time, tmp_priority, tmp_time_start)
        print (first.name, first.time, first.priority, first.time_start)
    elif sum_struct == 1:
        secound = Flow(tmp_name, tmp_time, tmp_priority, tmp_time_start)
        print (secound.name, secound.time, secound.priority, secound.time_start)
    elif sum_struct == 2:
        third = Flow(tmp_name, tmp_time, tmp_priority, tmp_time_start)
        print (third.name, third.time, third.priority, third.time_start)
    elif sum_struct == 3:
        fourth = Flow(tmp_name, tmp_time, tmp_priority, tmp_time_start)
        print (fourth.name, fourth.time, fourth.priority, fourth.time_start)
    sum_struct += 1

wb = openpyxl.load_workbook(filename = 'text.xlsx')
sheet = wb['Лист3']

#Заполнение таблицы процессов
sheet['A3'] = first.name
sheet['B3'] = first.time
sheet['C3'] = first.priority
sheet['D3'] = first.time_start

sheet['A4'] = secound.name
sheet['B4'] = secound.time
sheet['C4'] = secound.priority
sheet['D4'] = secound.time_start

sheet['A5'] = third.name
sheet['B5'] = third.time
sheet['C5'] = third.priority
sheet['D5'] = third.time_start

sheet['A6'] = fourth.name
sheet['B6'] = fourth.time
sheet['C6'] = fourth.priority
sheet['D6'] = fourth.time_start

#Создание массива квантов
all_time = first.time + secound.time + third.time + fourth.time
data = array('i', [])
for i in range(50):
    data.insert(i, 0)

#Алгоритм FCFS
min_start = first.time_start
flag = 0
if min_start > secound.time_start:
    min_start = secound.time_start
    flag = 1
if min_start > third.time_start:
    min_start = third.time_start
    flag = 2
if min_start > fourth.time_start:
    min_start = fourth.time_start
    flag = 3


if flag == 0:
    for j in range(int(first.time + first.time_start)):
        if j < first.time_start:
            continue
        sheet.cell(row =3, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
        data[j]+=1
        
    expecttion (data, secound)
    expecttion (data, third)
    expecttion (data, fourth)
    min_start = secound.time_start
    if min_start > third.time_start:
        min_start = third.time_start
        flag = 1
    if min_start > fourth.time_start:
        min_start = fourth.time_star
        flag = 2
    fill_cvant_1(flag, secound, third, fourth)

elif flag == 1:
    for j in range(int(secound.time + secound.time_start)):
        if j < secound.time_start:
            continue
        sheet.cell(row =4, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
        data[j]+=1
        
    expecttion (data, first)
    expecttion (data, third)
    expecttion (data, fourth)
    flag = 0
    min_start = first.time_start
    if min_start > thirt.time_start:
        min_start = thirt.time_start
        flag = 1
    if min_start > fourth.time_start:
        min_start = fourth.time_start
        flag = 2
    fill_cvant_1(flag, first, third, fourth)

elif flag == 2:
    for j in range(int(third.time + third.time_start)):
        if j < third.time_start:
            continue
        sheet.cell(row =5, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
        data[j]+=1
        
    expecttion (data, secound)
    expecttion (data, first)
    expecttion (data, fourth)
    flag = 0
    min_start = first.time_start
    if min_start > secound.time_start:
        min_start = secound.time_start
        flag = 1
    if min_start > fourth.time_start:
        min_start = fourth.time_start
        flag = 2
    fill_cvant_1(flag, first, secound, fourth)
elif flag == 3:
    for j in range(int(fourth.time + fourth.time_start)):
        if j < fourth.time_start:
            continue
        sheet.cell(row =6, column=j+4).fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
        data[j]+=1
        
    expecttion (data, secound)
    expecttion (data, third)
    expecttion (data, first)
    flag = 0
    min_start = first.time_start
    if min_start > secound.time_start:
        min_start = secound.time_start
        flag = 1
    if min_start > third.time_start:
        min_start = third.time_start
        flag = 2
    fill_cvant_1(flag, first, secound, third)



#Занесение конечных данных
sheet['AI3'] = first.time + first.time_w8
sheet['AJ3'] = first.time_w8
sheet['AK3'] = first.time/(first.time + first.time_w8)
sheet['AL3'] = (first.time + first.time_w8)/first.time

sheet['AI4'] = secound.time + secound.time_w8
sheet['AJ4'] = secound.time_w8
sheet['AK4'] = secound.time/(secound.time + secound.time_w8)
sheet['AL4'] = (secound.time + secound.time_w8)/secound.time

sheet['AI5'] = third.time + third.time_w8
sheet['AJ5'] = third.time_w8
sheet['AK5'] = third.time/(third.time + third.time_w8)
sheet['AL5'] = (third.time + third.time_w8)/third.time

sheet['AI6'] = fourth.time + fourth.time_w8
sheet['AJ6'] = fourth.time_w8
sheet['AK6'] = fourth.time/(fourth.time + fourth.time_w8)
sheet['AL6'] = (fourth.time + fourth.time_w8)/fourth.time

#Расчёт средних значений
sheet['AI7'] = (sheet['AI3'].value + sheet['AI4'].value + sheet['AI5'].value + sheet['AI6'].value)/4
sheet['AJ7'] = (sheet['AJ3'].value + sheet['AJ4'].value + sheet['AJ5'].value + sheet['AJ6'].value)/4
sheet['AK7'] = (sheet['AK3'].value + sheet['AK4'].value + sheet['AK5'].value + sheet['AK6'].value)/4
sheet['AL7'] = (sheet['AL3'].value + sheet['AL4'].value + sheet['AL5'].value + sheet['AL6'].value)/4

wb.save('text.xlsx')